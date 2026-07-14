import logging
import math
import json
from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Avg, Count, Q
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.contrib.auth import get_user_model
from decimal import Decimal, InvalidOperation
from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import user_passes_test
from django.views.decorators.http import require_http_methods
from users.decorators import get_client_ip
from audit.models import AuditLog          # ← الموقع الصحيح للـ AuditLog
from .forms import WarehouseForm, StationForm 
# ملاحظة: Warehouse و get_object_or_404 و messages و Q و redirect مستوردة لديك مسبقاً
from users.security import is_trusted_device
from users.decorators import role_required
from .models import Trip, DischargeRecord, Station, Warehouse,Truck, Driver
from .forms import DischargeForm, FaultReportForm
from faults.models import MechanicalFault
from alerts.models import Alert
from django.urls import reverse
from .utils import generate_qr_code_base64 
from .forms import DischargeForm, FaultReportForm,DriverForm,TruckForm
from geopy.distance import geodesic
from django.urls import reverse
import re
from django.db import IntegrityError

def _normalize_seals(raw):
    """يحوّل نص أرقام الأختام إلى مجموعة موحّدة.
    يتجاهل: الترتيب، المسافات الزائدة، حالة الأحرف، ونوع الفاصل (، , ; مسافة سطر)."""
    if not raw:
        return set()
    parts = re.split(r'[,\u060C;\n\r\t ]+', raw.strip())
    return {p.strip().upper() for p in parts if p.strip()}

User = get_user_model()
logger = logging.getLogger(__name__)

# ============================================================================
# 🧮 INTERNAL HELPERS - الدوال المساعدة الداخلية
# ============================================================================

def _calculate_haversine(lat1, lon1, lat2, lon2):
    """حساب المسافة بين نقطتين جغرافيتين بوحدة المتر باستخدام صيغة Haversine."""
    R = 6371000  # نصف قطر الأرض بالمتر
    phi1 = math.radians(float(lat1))
    phi2 = math.radians(float(lat2))
    delta_phi = math.radians(float(lat2) - float(lat1))
    delta_lambda = math.radians(float(lon2) - float(lon1))

    a = math.sin(delta_phi / 2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


def _verify_qr_scanned(trip, request):
    """تحقق من أن التفريغ يتم بناءً على مسح حديث للـ QR (صلاحية نافذة التفريغ 6 ساعات)."""
    if not trip.qr_scanned_at:
        logger.warning(f'🚫 محاولة تفريغ مباشر بالروابط دون مسح QR للرحلة: {trip.trip_code}')
        raise PermissionDenied('يجب مسح رمز QR الفعلي للشاحنة أولاً قبل إدخال البيانات!')
    
    time_diff = timezone.now() - trip.qr_scanned_at
    if time_diff.total_seconds() > 21600:  # 6 ساعات كحد أقصى للتفريغ بعد المسح
        logger.warning(f'⚠️ انتهت نافذة التفريغ المسموحة بعد المسح للرحلة: {trip.trip_code}')
        raise PermissionDenied('مضى وقت طويل (أكثر من 6 ساعات) منذ مسح QR. يرجى إعادة مسح الرمز عند الخزان.')


# ============================================================================
# 📋 TRIPS MANAGEMENT - إدارة وعرض الرحلات
# ============================================================================
def _is_admin(user):
    """شرط الصلاحية: مستخدم مسجّل ودوره admin."""
    return user.is_authenticated and getattr(user, 'role', None) == 'admin'


def get_base_template_for(user):
    """اختيار القالب الأساسي (الشريط الجانبي) حسب دور المستخدم."""
    return {
        'dispatcher': 'dispatcher_base.html',
        'agent':      'agent_base.html',
        'admin':      'admin_base.html',
        'manager':    'manager_base.html',
        'auditor':    'auditor_base.html',
    }.get(getattr(user, 'role', None), 'dispatcher_base.html') 
    
@login_required
@role_required('dispatcher', 'agent', 'manager', 'auditor', 'admin')
def trip_list(request):
    """عرض قائمة الرحلات مفلترة حسب صلاحيات ودور المستخدم الحالي."""
    user = request.user
    trips = Trip.objects.select_related('truck', 'station', 'warehouse', 'dispatcher', 'agent').order_by('-created_at')

    if user.role == 'dispatcher':
        trips = trips.filter(dispatcher=user)
    elif user.role == 'agent':
        trips = trips.filter(agent=user)
    elif user.role == 'auditor':
        trips = trips.filter(status__in=['completed', 'suspect'])

    # هذا هو الجزء الحرج الذي يغذي القوائم المنسدلة في الواجهة بالبيانات
    context = {
        'trips': trips,
        'can_create': user.role in ['dispatcher', 'admin'],
        'stations': Station.objects.filter(is_active=True),
        'trucks': Truck.objects.filter(status='active'),
        'agents': User.objects.filter(role='agent', is_active=True),
        'base_template': get_base_template_for(request.user),
    }
    return render(request, 'trips/trip_list.html', context)

@login_required
@role_required('dispatcher', 'agent', 'manager', 'auditor', 'admin')
def trip_detail(request, trip_id):
    """عرض التفاصيل العميقة لرحلة معينة مع سجل التفريغ إن وجد."""
    trip = get_object_or_404(
        Trip.objects.select_related('truck', "truck__driver", 'station', 'warehouse', 'dispatcher', 'agent'),
        pk=trip_id
    )

    # فحص صلاحية الوصول للرحلة بناءً على الدور
    if request.user.role == 'dispatcher' and trip.dispatcher != request.user:
        raise PermissionDenied()
    if request.user.role == 'agent' and trip.agent != request.user:
        raise PermissionDenied()

    discharge_record = getattr(trip, 'discharge_record', None)

    # ✅ توليد صورة QR (فقط إذا كانت الرحلة نشطة والرمز صالح)
    qr_image = None
    if trip.is_qr_valid() and trip.status == 'pending':
        qr_url = request.build_absolute_uri(
            reverse('trips:qr_redirect', args=[trip.qr_token])
        )
        qr_image = generate_qr_code_base64(qr_url)

    context = {
        'trip': trip,
        'discharge_record': discharge_record,
        'qr_image': qr_image,   # ← أضفناه للسياق
        'base_template': get_base_template_for(request.user),
        
    }
    return render(request, 'trips/trip_detail.html', context)



@login_required
@require_http_methods(["GET", "POST"])  # التعديل الحاسم: تم استبدال الديكوريتور المكسور بالديكوريتور القياسي المدمج لدجانجو
@role_required('dispatcher', 'admin')
@transaction.atomic
def create_trip(request):
    """إنشاء رحلة جديدة وتوليد بيانات التتبع والـ QR تلقائياً."""
    if request.method == 'POST':
        try:
            agent_id = request.POST.get('agent_id')
            truck_id = request.POST.get('truck_id')
            station_id = request.POST.get('station_id')
            seal_numbers = request.POST.get('seal_numbers', '').strip()

            shipped_volume = Decimal(request.POST.get('shipped_volume_ambient', 0))
            shipped_temp = Decimal(request.POST.get('shipped_temperature', 0))
            shipped_density = Decimal(request.POST.get('shipped_density', 0))
            notes = request.POST.get('notes', '').strip()

            if not all([agent_id, truck_id, station_id, seal_numbers]):
                messages.error(request, 'يرجى ملء جميع الحقول المطلوبة.')
                return redirect('trips:create_trip')

            if shipped_volume <= 0 or shipped_temp < -50 or shipped_temp > 60 \
                or shipped_density < Decimal('0.653') or shipped_density > Decimal('0.770'):
                messages.error(request, 'البيانات الفيزيائية غير منطقية للبنزين (الكثافة يجب أن تكون بين 0.653 و 0.770).')
                return redirect('trips:create_trip')

            agent = get_object_or_404(User, id=agent_id, role='agent')
            truck = get_object_or_404(Truck, id=truck_id)
            station = get_object_or_404(Station, id=station_id)
            # ⭐ منع إسناد محطة غير مرتبطة بالمندوب (حماية من تجاوز فلترة الواجهة)
            if not station.agents.filter(id=agent.id).exists():
                messages.error(request, 'المحطة المختارة غير مرتبطة بهذا المندوب. اختر محطة من محطاته المعينة.')
                return redirect('trips:create_trip')

            # ⭐ (1) قفل صف السائق حتى نهاية المعاملة — يسلسل الطلبات المتزامنة (ضد Race Condition)
            driver = Driver.objects.select_for_update().get(pk=truck.driver_id)
            if driver.has_active_trip():
                messages.error(request, f'السائق {driver.driver_name} لديه رحلة نشطة لم تُفرَّغ بعد. لا يمكن إسناده لرحلة ثانية.')
                return redirect('trips:create_trip')
            
            # جلب مستودع الموظف الحالي
            warehouse = request.user.assigned_warehouses.first()
            if not warehouse:
                messages.error(request, 'خطأ: حسابك غير مرتبط بأي مستودع لإنشاء رحلات منه.')
                return redirect('trips:trip_list')

            trip = Trip(
                dispatcher=request.user,
                agent=agent,
                truck=truck,
                station=station,
                warehouse=warehouse,
                seal_numbers=seal_numbers,
                shipped_volume_ambient=shipped_volume,
                shipped_temperature=shipped_temp,
                shipped_density=shipped_density,
                notes=notes,
                status='pending'
            )
            try:
                trip.save()
            except IntegrityError:
                messages.error(request, 'تعذّر الإنشاء: هذه الشاحنة مُسنَدة لرحلة نشطة بالفعل (طلب متزامن).')
                return redirect('trips:create_trip')

            logger.info(f'✅ {request.user.username} أنشأ الشحنة: {trip.trip_code}')
            messages.success(request, f'تم إنشاء الرحلة {trip.trip_code} وتوليد الـ QR بنجاح.')
            return redirect('trips:trip_detail', trip_id=trip.id)

        except (ValueError, InvalidOperation):
            messages.error(request, 'خطأ في معالجة البيانات الرقمية المدخلة.')
            return redirect('trips:create_trip')

    # ============ GET: تجهيز القوائم ============
    # ⭐ (2) استبعاد شاحنات السائقين المشغولين + حصرها بمستودع الموظف
    warehouse = request.user.assigned_warehouses.first()

    busy_driver_ids = Trip.objects.filter(
        status__in=['pending', 'in_transit']
    ).values_list('truck__driver_id', flat=True)

    available_trucks = Truck.objects.filter(status='active').exclude(driver_id__in=busy_driver_ids)
    if warehouse:
        available_trucks = available_trucks.filter(driver__warehouse=warehouse)

    context = {
        'agents': User.objects.filter(role='agent', is_active=True),
        'trucks': available_trucks,
        'stations': Station.objects.filter(is_active=True, is_banned=False).prefetch_related('agents'),
        'base_template': get_base_template_for(request.user),
    }
    return render(request, 'trips/create_trip.html', context)

@login_required
@role_required('dispatcher', 'manager', 'admin')
def update_trip(request, trip_id):
    """تعديل بيانات الرحلة ما دامت في الحالة المبدئية."""
    trip = get_object_or_404(Trip, pk=trip_id)
    if trip.status != 'pending':
        messages.error(request, 'لا يمكن تعديل بيانات شحنة انطلقت أو أغلقت بالفعل.')
        return redirect('trips:trip_detail', trip_id=trip.id)
    return render(request, 'trips/update_trip.html', {'trip': trip})

@login_required
@require_http_methods(["POST"])
@role_required('dispatcher', 'manager', 'admin')
def mark_trip_departed(request, trip_id):
    """إدارة الحالة: تحديث حالة الشحنة إلى 'في الطريق' إيذاناً بانطلاقها ميدانياً."""
    trip = get_object_or_404(Trip, pk=trip_id)
    if request.user.role == 'dispatcher' and trip.dispatcher != request.user:
        raise PermissionDenied()
    if trip.status != 'pending':
        messages.error(request, 'لا يمكن تحديث الحالة: الشحنة ليست في وضع الانتظار.')
        return redirect('trips:trip_detail', trip_id=trip.id)
    trip.status = 'in_transit'
    trip.save(update_fields=['status', 'updated_at'])
    AuditLog.objects.create(
        user=request.user, action='mark_trip_departed',
        target_table='shield_trips', target_id=trip.id,
        old_value='pending', new_value='in_transit',
        ip_address=get_client_ip(request),
    )
    messages.success(request, f'🚚 تم إطلاق الشحنة {trip.trip_code} وتحديث حالتها إلى «في الطريق».')
    return redirect('trips:trip_detail', trip_id=trip.id)


@login_required
@role_required('auditor', 'manager', 'admin')
def shipments_map(request):
    """الخريطة التفاعلية لمتابعة حركة الشحنات ومواقع التفريغ الفعلية."""
    trips = (
        Trip.objects
        .select_related('station', 'truck', 'truck__driver', 'agent')
        .exclude(status='canceled')
        .order_by('-created_at')[:200]
    )
    stations = Station.objects.filter(is_active=True)

    station_points = [{
        'lat': float(s.latitude), 'lng': float(s.longitude),
        'name': s.station_name, 'customer': s.customer_number,
        'city': s.city, 'radius': s.geofence_radius,
    } for s in stations]

    discharge_points = []
    active_points = []
    for trip in trips:
        rec = getattr(trip, 'discharge_record', None)
        agent_name = (trip.agent.full_name or trip.agent.username) if trip.agent else '—'
        if rec and rec.scan_latitude and rec.scan_longitude:
            discharge_points.append({
                'lat': float(rec.scan_latitude), 'lng': float(rec.scan_longitude),
                'trip_code': trip.trip_code, 'station': trip.station.station_name,
                'agent': agent_name, 'truck': trip.truck.truck_id,
                'status': trip.status, 'status_display': trip.get_status_display(),
                'geofence': rec.geofence_status, 'distance': float(rec.scan_distance_m),
                'is_suspect': rec.is_suspect,
                'station_lat': float(trip.station.latitude),
                'station_lng': float(trip.station.longitude),
            })
        elif trip.status in ('pending', 'in_transit'):
            active_points.append({
                'lat': float(trip.station.latitude), 'lng': float(trip.station.longitude),
                'trip_code': trip.trip_code, 'station': trip.station.station_name,
                'agent': agent_name, 'truck': trip.truck.truck_id,
                'status': trip.status, 'status_display': trip.get_status_display(),
            })

    context = {
        'station_points': station_points,
        'discharge_points': discharge_points,
        'active_points': active_points,
        'total_active': len(active_points),
        'total_discharge': len(discharge_points),
        'total_suspect': sum(1 for p in discharge_points if p['is_suspect']),
        'base_template':get_base_template_for(request.user),
    }
    return render(request, 'trips/shipments_map.html', context)
@login_required
@role_required('manager', 'admin')
def delete_trip(request, trip_id):
    """حذف الرحلة نهائياً من النظام (للمديرين فقط)."""
    trip = get_object_or_404(Trip, pk=trip_id)
    trip_code = trip.trip_code
    trip.delete()
    messages.success(request, f'تم حذف الشحنة {trip_code} نهائياً من المنظومة.')
    return redirect('trips:trip_list')


# ============================================================================
# 🔐 QR SECURITY - بوابة مسح الرموز الأمنية
# ============================================================================

@login_required
@role_required('agent', 'dispatcher', 'manager', 'admin')
def qr_redirect(request, qr_token):
    """مسح الـ QR Code والتحقق من الهوية والموقع لمنع هجمات التخمين."""
    try:
        trip = Trip.objects.get(qr_token=qr_token)
        
        if trip.status in ['completed', 'suspect', 'canceled']:
            messages.error(request, '🚫 هذه الرحلة مغلقة أو تم تفريغها مسبقاً.')
            return redirect('trips:trip_list')
            
        if not trip.is_qr_valid():
            logger.warning(f'🚫 محاولة مسح QR منتهي الصلاحية الزمني للرحلة: {trip.trip_code}')
            messages.error(request, '⏰ انتهت الصلاحية الزمنية لرمز QR الحالي.')
            return redirect('trips:trip_list')
        
        if request.user.role == 'agent' and trip.agent != request.user:
            logger.warning(f'🚨 اختراق صلاحيات: {request.user.username} مسح QR لرحلة غير مخصصة له.')
            messages.error(request, '❌ رمز الـ QR هذا ليس مخصصاً لحسابك الميداني.')
            return redirect('trips:trip_list')
        
        trip.qr_scanned_at = timezone.now()
        trip.save()
        
        logger.info(f'✅ تم مسح QR بنجاح للشحنة {trip.trip_code} بواسطة {request.user.username}')
        messages.success(request, f'✓ تم التحقق من الكود المطبوع. يمكنك الآن تعبئة نموذج التفريغ للشحنة: {trip.trip_code}')
        return redirect('trips:trip_detail', trip_id=trip.id)
        
    except Trip.DoesNotExist:
        logger.error(f'🚨 محاولة تخمين أو مسح توكن QR وهمي: {qr_token} من المستخدم {request.user.username}')
        messages.error(request, '❌ رمز QR غير معتمد أو غير موجود بقاعدة البيانات.')
        return redirect('trips:trip_list')


# ============================================================================
# 📥 DISCHARGE WORKFLOW - معالجة عمليات التفريغ والجيوفينسنج
# ============================================================================


@login_required
@role_required('dispatcher', 'agent', 'manager', 'auditor', 'admin')
def discharge_records(request):
    """عرض سجلات التاريخية لعمليات التفريغ المقفلة."""
    records = DischargeRecord.objects.select_related('trip', 'agent').order_by('-created_at')
    return render(request, 'trips/discharge_records.html', {'records': records})
@user_passes_test(_is_admin, login_url='users:login')
def truck_list(request):
    qs = Truck.objects.select_related('driver').order_by('truck_id')
    search = request.GET.get('q', '').strip()
    if search:
        qs = qs.filter(Q(truck_id__icontains=search) | Q(driver__driver_name__icontains=search))
    return render(request, 'trips/truck_list.html', {'trucks': qs, 'search': search})





# ============================================================================
# 🚨 TAMPER DETECTION & AUDITING - الرصد والتدقيق الأمنى
# ============================================================================

@login_required
@role_required('dispatcher', 'agent', 'manager', 'auditor', 'admin')
def suspect_trips(request):
    """لوحة حصر وتحليل الشحنات المثيرة للاشتباه (الخرق الجغرافي، كسر الأختام، العجز الحجمي)."""
    suspect_records = DischargeRecord.objects.filter(is_suspect=True).select_related(
        'trip', 'trip__truck', 'trip__station', 'trip__warehouse', 'agent'
    ).order_by('-created_at')
    
    if request.user.role == 'dispatcher':
        suspect_records = suspect_records.filter(trip__dispatcher=request.user)
    elif request.user.role == 'agent':
        suspect_records = suspect_records.filter(trip__agent=request.user)
    
    context = {
        'suspect_records': suspect_records,
        'total_suspect': suspect_records.count(),
        'geofence_suspect': suspect_records.filter(geofence_status='red').count(),
        'seal_suspect': suspect_records.filter(seal_check_passed=False).count(),
    }
    return render(request, 'trips/suspect_trips.html', context)


# ============================================================================
# 📊 REPORTS & ANALYTICS - التقارير الإحصائية المجمعة
# ============================================================================

@login_required
@role_required('dispatcher', 'manager', 'auditor', 'admin')
def trip_reports(request):
    """لوحة قيادة مجمعة للتحليلات الإحصائية ونسب الفقد العام + تصدير PDF/Excel."""
    avg_variance = DischargeRecord.objects.aggregate(avg=Avg('variance_percent'))['avg'] or 0

    context = {
        'total_trips':      Trip.objects.count(),
        'completed_trips':  Trip.objects.filter(status='completed').count(),
        'suspect_trips':    Trip.objects.filter(status='suspect').count(),
        'pending_trips':    Trip.objects.filter(status='pending').count(),
        'in_transit_trips': Trip.objects.filter(status='in_transit').count(),
        'avg_variance':     round(avg_variance, 3),
        'worst_trips':      DischargeRecord.objects.select_related(
                                'trip', 'trip__station', 'agent'
                            ).order_by('-variance_percent')[:5],
        'agent_performance': DischargeRecord.objects.values('agent__username').annotate(
                                total=Count('id'),
                                suspect_count=Count('id', filter=Q(is_suspect=True))
                            ).order_by('-suspect_count')[:10],
    }

        # 📗 تصدير Excel احترافي (.xlsx بألوان وحدود واتجاه RTL)
    if request.GET.get('export') == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير Shield"
        ws.sheet_view.rightToLeft = True

        # الأنماط
        title_font   = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
        header_font  = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        label_font   = Font(name='Calibri', size=11, bold=True)
        normal_font  = Font(name='Calibri', size=11)
        center = Alignment(horizontal='center', vertical='center')
        right  = Alignment(horizontal='right', vertical='center')
        thin   = Side(style='thin', color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_fill   = PatternFill('solid', fgColor="1E3A8A")   # أزرق داكن
        header_fill  = PatternFill('solid', fgColor="059669")   # أخضر
        section_fill = PatternFill('solid', fgColor="334155")   # رمادي

        def style_row(row_idx, c1, c2, font=None, fill=None, align=None, bordered=True):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row=row_idx, column=c)
                if font:  cell.font = font
                if fill:  cell.fill = fill
                if align: cell.alignment = align
                if bordered: cell.border = border

        r = 1
        # العنوان
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(row=r, column=1, value="تقرير Shield — العجز والفقد وتحليل الأداء")
        style_row(r, 1, 5, font=title_font, fill=title_fill, align=center, bordered=False)
        ws.row_dimensions[r].height = 30
        r += 2

        # المؤشرات
        ws.cell(row=r, column=1, value="المؤشر"); ws.cell(row=r, column=2, value="القيمة")
        style_row(r, 1, 2, font=header_font, fill=header_fill, align=center)
        r += 1
        for label, val in [
            ("إجمالي الشحنات", context['total_trips']),
            ("مكتملة", context['completed_trips']),
            ("في الطريق", context['in_transit_trips']),
            ("قيد الانتظار", context['pending_trips']),
            ("مشبوهة", context['suspect_trips']),
            ("متوسط العجز %", context['avg_variance']),
        ]:
            ws.cell(row=r, column=1, value=label).font = label_font
            ws.cell(row=r, column=2, value=val).font = normal_font
            style_row(r, 1, 2, align=center)
            ws.cell(row=r, column=1).alignment = right
            r += 1
        r += 1

        # أعلى الشحنات فقداً
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(row=r, column=1, value="أعلى الشحنات فقداً")
        style_row(r, 1, 5, font=header_font, fill=section_fill, align=center)
        r += 1
        for i, h in enumerate(["رمز الشحنة", "المحطة", "المندوب", "العجز (لتر)", "النسبة %"], start=1):
            ws.cell(row=r, column=i, value=h)
        style_row(r, 1, 5, font=header_font, fill=header_fill, align=center)
        r += 1
        for rec in context['worst_trips']:
            agent = rec.agent.full_name or rec.agent.username
            ws.cell(row=r, column=1, value=rec.trip.trip_code)
            ws.cell(row=r, column=2, value=rec.trip.station.station_name)
            ws.cell(row=r, column=3, value=agent)
            ws.cell(row=r, column=4, value=float(rec.variance_liters))
            ws.cell(row=r, column=5, value=float(rec.variance_percent))
            style_row(r, 1, 5, font=normal_font, align=center)
            r += 1
        r += 1

        # أداء المندوبين
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.cell(row=r, column=1, value="أداء المندوبين")
        style_row(r, 1, 3, font=header_font, fill=section_fill, align=center)
        r += 1
        for i, h in enumerate(["المندوب", "إجمالي التفريغ", "المشبوهة"], start=1):
            ws.cell(row=r, column=i, value=h)
        style_row(r, 1, 3, font=header_font, fill=header_fill, align=center)
        r += 1
        for row in context['agent_performance']:
            ws.cell(row=r, column=1, value=row['agent__username'])
            ws.cell(row=r, column=2, value=row['total'])
            ws.cell(row=r, column=3, value=row['suspect_count'])
            style_row(r, 1, 3, font=normal_font, align=center)
            r += 1

        # عرض الأعمدة
        for i, w in enumerate([22, 26, 26, 14, 12], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="shield_report.xlsx"'
        wb.save(response)
        return response

    return render(request, 'trips/reports.html', context)


@login_required
@role_required('dispatcher', 'manager', 'auditor', 'admin')
def trip_report_detail(request, trip_id):
    """إنتاج تقرير تدقيق هندسي مفصل وشامل لرحلة توريد فردية."""
    trip = get_object_or_404(Trip.objects.prefetch_related('discharge_record'), pk=trip_id)

    # 🔒 منع IDOR: موظف المستودع يرى تقارير رحلاته فقط (المدير/المراقب/الأدمن يرون الكل)
    if request.user.role == 'dispatcher' and trip.dispatcher != request.user:
        raise PermissionDenied()

    return render(request, 'trips/trip_report_detail.html', {
        'trip': trip,
        'discharge_record': getattr(trip, 'discharge_record', None),
    })


# ============================================================================
# 🏭 WAREHOUSE MASTER-DATA MANAGEMENT (admin فقط)
# ============================================================================



@user_passes_test(_is_admin, login_url='users:login')
def warehouse_list(request):
    """قراءة (Read): عرض المستودعات + بحث."""
    qs = Warehouse.objects.all().order_by('warehouse_name')
    search = request.GET.get('q', '').strip()
    if search:
        qs = qs.filter(
            Q(warehouse_code__icontains=search) |
            Q(warehouse_name__icontains=search) |
            Q(location_city__icontains=search)
        )
    return render(request, 'trips/warehouse_list.html', {
        'warehouses': qs, 'search': search,
    })


@user_passes_test(_is_admin, login_url='users:login')
def warehouse_create(request):
    """إنشاء (Create) + تسجيل تدقيق."""
    if request.method == 'POST':
        form = WarehouseForm(request.POST)
        if form.is_valid():
            wh = form.save()
            AuditLog.objects.create(
                user=request.user, action='create_warehouse',
                target_table='warehouses', target_id=wh.id,
                new_value=f'{wh.warehouse_code} - {wh.warehouse_name}',
                ip_address=get_client_ip(request),
            )
            messages.success(request, f'✅ تم إنشاء المستودع «{wh.warehouse_name}» بنجاح.')
            return redirect('trips:warehouse_list')
    else:
        form = WarehouseForm()
    return render(request, 'trips/warehouse_form.html', {'form': form, 'mode': 'create'})


@user_passes_test(_is_admin, login_url='users:login')
def warehouse_update(request, pk):
    """تعديل (Update) + تسجيل القيمة قبل/بعد."""
    wh = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        old_value = f'{wh.warehouse_code} - {wh.warehouse_name} | نشط={wh.is_active}'
        form = WarehouseForm(request.POST, instance=wh)
        if form.is_valid():
            wh = form.save()
            new_value = f'{wh.warehouse_code} - {wh.warehouse_name} | نشط={wh.is_active}'
            AuditLog.objects.create(
                user=request.user, action='update_warehouse',
                target_table='warehouses', target_id=wh.id,
                old_value=old_value, new_value=new_value,
                ip_address=get_client_ip(request),
            )
            messages.success(request, f'✅ تم تحديث المستودع «{wh.warehouse_name}» بنجاح.')
            return redirect('trips:warehouse_list')
    else:
        form = WarehouseForm(instance=wh)
    return render(request, 'trips/warehouse_form.html', {'form': form, 'mode': 'edit', 'warehouse': wh})


@user_passes_test(_is_admin, login_url='users:login')
@require_http_methods(["POST"])
def warehouse_archive(request, pk):
    """أرشفة/استعادة المستودع (Soft Delete) — لا حذف فعلي، فقط تغيير is_active."""
    wh = get_object_or_404(Warehouse, pk=pk)

    old = 'active' if wh.is_active else 'archived'
    wh.is_active = not wh.is_active
    wh.save(update_fields=['is_active'])
    new = 'active' if wh.is_active else 'archived'

    action = 'restore_warehouse' if wh.is_active else 'archive_warehouse'
    AuditLog.objects.create(
        user=request.user, action=action,
        target_table='warehouses', target_id=wh.id,
        old_value=old, new_value=new, ip_address=get_client_ip(request),
    )
    verb = 'استعادة' if wh.is_active else 'أرشفة'
    messages.success(request, f'✅ تم {verb} المستودع «{wh.warehouse_name}».')
    return redirect('trips:warehouse_list')

# ============================================================================
# ⛽ STATION MASTER-DATA MANAGEMENT (admin فقط)
# ============================================================================

@user_passes_test(_is_admin, login_url='users:login')
def stations_list(request):
    """قراءة: عرض المحطات + بحث."""
    qs = Station.objects.all().order_by('station_name')
    search = request.GET.get('q', '').strip()
    if search:
        qs = qs.filter(
            Q(customer_number__icontains=search) |
            Q(station_name__icontains=search) |
            Q(city__icontains=search)
        )
    return render(request, 'trips/station_list.html', {'stations': qs, 'search': search})


@user_passes_test(_is_admin, login_url='users:login')
def station_create(request):
    """إنشاء + تدقيق."""
    if request.method == 'POST':
        form = StationForm(request.POST)
        if form.is_valid():
            st = form.save()
            AuditLog.objects.create(
                user=request.user, action='create_station',
                target_table='stations', target_id=st.id,
                new_value=f'{st.customer_number} - {st.station_name}',
                ip_address=get_client_ip(request),
            )
            messages.success(request, f'✅ تم إنشاء المحطة «{st.station_name}» بنجاح.')
            return redirect('trips:stations_list')
    else:
        form = StationForm()
    return render(request, 'trips/station_form.html', {'form': form, 'mode': 'create'})


@user_passes_test(_is_admin, login_url='users:login')
def station_update(request, pk):
    """تعديل + تدقيق قبل/بعد."""
    st = get_object_or_404(Station, pk=pk)
    if request.method == 'POST':
        old_value = f'{st.customer_number} - {st.station_name} | نشط={st.is_active} موقوفة={st.is_banned}'
        form = StationForm(request.POST, instance=st)
        if form.is_valid():
            st = form.save()
            new_value = f'{st.customer_number} - {st.station_name} | نشط={st.is_active} موقوفة={st.is_banned}'
            AuditLog.objects.create(
                user=request.user, action='update_station',
                target_table='stations', target_id=st.id,
                old_value=old_value, new_value=new_value,
                ip_address=get_client_ip(request),
            )
            messages.success(request, f'✅ تم تحديث المحطة «{st.station_name}» بنجاح.')
            return redirect('trips:stations_list')
    else:
        form = StationForm(instance=st)
    return render(request, 'trips/station_form.html', {'form': form, 'mode': 'edit', 'station': st})


@user_passes_test(_is_admin, login_url='users:login')
@require_http_methods(["POST"])
def station_archive(request, pk):
    """أرشفة/استعادة المحطة (Soft Delete) — لا حذف فعلي، فقط تغيير is_active."""
    st = get_object_or_404(Station, pk=pk)

    old = 'active' if st.is_active else 'archived'
    st.is_active = not st.is_active
    st.save(update_fields=['is_active'])
    new = 'active' if st.is_active else 'archived'

    action = 'restore_station' if st.is_active else 'archive_station'
    AuditLog.objects.create(
        user=request.user, action=action,
        target_table='stations', target_id=st.id,
        old_value=old, new_value=new, ip_address=get_client_ip(request),
    )
    verb = 'استعادة' if st.is_active else 'أرشفة'
    messages.success(request, f'✅ تم {verb} المحطة «{st.station_name}».')
    return redirect('trips:stations_list')
# ============================================================================
# 🚉 STATION AGENT VIEWS — واجهات مندوب المحطة
# ============================================================================

def _get_client_ip(request):
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded:
        return x_forwarded.split(',')[0]
    return request.META.get('REMOTE_ADDR', '')


def _require_agent(request):
    return request.user.is_authenticated and request.user.role == 'agent'


@login_required
def agent_trip_list(request):
    """الرحلات القادمة المسندة إلى المندوب الحالي والتي لم تُغلق بعد."""
    if not _require_agent(request):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة.')
        return redirect('dashboard:home')
    trips = (
        Trip.objects.filter(agent=request.user)
        .exclude(status__in=['completed', 'canceled', 'suspect'])
        .select_related('station', 'truck', 'truck__driver', 'warehouse')
        .order_by('created_at')
    )
    return render(request, 'trips/agent_trip_list.html', {'trips': trips})


@login_required
def agent_discharge_log(request):
    """سجل التفريغ الكامل لعمليات التفريغ التي وثّقها المندوب الحالي."""
    if not _require_agent(request):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة.')
        return redirect('dashboard:home')
    records = (
        DischargeRecord.objects.filter(agent=request.user)
        .select_related('trip', 'trip__station', 'trip__truck')
        .order_by('-created_at')
    )
    return render(request, 'trips/discharge_log.html', {'records': records})


@login_required
def agent_fault_list(request):
    """إدارة البلاغات: عرض جميع البلاغات التي رفعها المندوب الحالي."""
    if not _require_agent(request):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة.')
        return redirect('dashboard:home')
    faults = (
        MechanicalFault.objects.filter(reported_by=request.user)
        .select_related('trip', 'trip__station', 'truck')
        .order_by('-created_at')
    )
    return render(request, 'trips/agent_fault_list.html', {'faults': faults})

@login_required
def agent_fault_detail(request, pk):
    """تفاصيل بلاغ رفعه المندوب الحالي — عرض كامل قابل للطباعة."""
    if not _require_agent(request):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة.')
        return redirect('dashboard:home')
    fault = get_object_or_404(
        MechanicalFault.objects.select_related('trip', 'trip__station', 'station', 'truck'),
        pk=pk, reported_by=request.user,   # 🔒 يمنع IDOR: بلاغاته هو فقط
    )
    return render(request, 'trips/agent_fault_detail.html', {'fault': fault})

@login_required
def submit_discharge(request, trip_id):
    """
    التوثيق الميداني والمطابقة الرقمية والجغرافية لتفريغ شحنة بنزين.
    الترتيب: أمان الجهاز → مطابقة QR → السياج الجغرافي → مطابقة الأختام → تسجيل وإغلاق الرحلة.
    """
    if not _require_agent(request):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة.')
        return redirect('dashboard:home')

    trip = get_object_or_404(Trip, pk=trip_id, agent=request.user)
    ip = _get_client_ip(request)

    if trip.status in ('completed', 'suspect', 'canceled'):
        messages.error(request, 'هذه الرحلة مغلقة بالفعل ولا يمكن تسجيل تفريغ جديد لها.')
        return redirect('trips:agent_trip_list')

    if request.method != 'POST':
        form = DischargeForm(trip=trip)
        return render(request, 'trips/discharge_form.html', {'trip': trip, 'form': form})

    # 1) فحص أمان الجهاز
    if not is_trusted_device(request):
        AuditLog.objects.create(
            user=request.user, action='untrusted_device_breach',
            target_table='shield_trips', target_id=trip.id, ip_address=ip,
            new_value=f'محاولة تسجيل تفريغ للرحلة {trip.trip_code} من جهاز غير موثّق.',
        )
        Alert.objects.create(
            trip=trip, alert_type='unauthorized_device', severity='critical',
            description=f'محاولة تفريغ للرحلة {trip.trip_code} من جهاز غير معتمد للمستخدم {request.user.username}. رُفضت العملية.',
        )
        messages.error(request, 'تم رفض العملية: هذا الجهاز غير موثوق أمنياً. تواصل مع مدير النظام لاعتماد جهازك.')
        return redirect('trips:submit_discharge', trip_id=trip.id)

    form = DischargeForm(request.POST, trip=trip)
    if not form.is_valid():
        return render(request, 'trips/discharge_form.html', {'trip': trip, 'form': form})

    # 2) المطابقة الرقمية (QR)
    scanned_token = request.POST.get('scanned_qr_token', '').strip()
    if not trip.is_qr_valid() or scanned_token != str(trip.qr_token):
        AuditLog.objects.create(
            user=request.user, action='qr_mismatch_breach',
            target_table='shield_trips', target_id=trip.id, ip_address=ip,
            new_value=f'توكن ممسوح غير مطابق أو منتهي الصلاحية للرحلة {trip.trip_code}.',
        )
        Alert.objects.create(
            trip=trip, alert_type='seal_tampering', severity='critical',
            description=f'فشل التحقق من QR للرحلة {trip.trip_code} من المندوب {request.user.username}. التوكن غير مطابق أو منتهٍ.',
        )
        messages.error(request, 'فشلت المطابقة الرقمية: رمز الـ QR غير مطابق لهذه الرحلة أو منتهي الصلاحية.')
        return redirect('trips:submit_discharge', trip_id=trip.id)

    # 3) السياج الجغرافي (Geofencing)
    station = trip.station
    agent_point = (float(form.cleaned_data['scan_latitude']), float(form.cleaned_data['scan_longitude']))
    station_point = (float(station.latitude), float(station.longitude))
    distance_m = geodesic(agent_point, station_point).meters

    if distance_m > station.geofence_radius:
        trip.status = 'suspect'
        trip.save(update_fields=['status', 'updated_at'])
        Alert.objects.create(
            trip=trip, alert_type='geofence_violation', severity='critical',
            description=f'تجاوز المندوب {request.user.username} النطاق ({station.geofence_radius}م) بمسافة {distance_m:.2f}م عن محطة {station.station_name}. عُلّقت الرحلة {trip.trip_code} كمشبوهة.',
        )
        AuditLog.objects.create(
            user=request.user, action='geofence_violation',
            target_table='shield_trips', target_id=trip.id, ip_address=ip,
            new_value=f'المسافة: {distance_m:.2f}م / المسموح: {station.geofence_radius}م',
        )
        messages.error(request, f'تم رفض العملية: أنت تبعد {distance_m:.2f}م عن المحطة، خارج نطاق الأمان ({station.geofence_radius}م). صُنّفت الرحلة كمشبوهة وأُخطر المراقبون.')
        return redirect('trips:agent_trip_list')

    # 4) نجاح — تسجيل التفريغ وإغلاق الرحلة
    record = form.save(commit=False)
    record.trip = trip
    record.agent = request.user
    record.scan_distance_m = round(distance_m, 2)
    record.geofence_status = 'green'

    # 🔒 مطابقة أرقام الأختام (غير حسّاسة للترتيب) — تُحسب في الخادم لا يقرّرها المندوب
    recorded_seals = _normalize_seals(trip.seal_numbers)
    observed_input = form.cleaned_data.get('observed_seal_numbers', '')
    observed_seals = _normalize_seals(observed_input)
    record.seal_check_passed = bool(recorded_seals) and (observed_seals == recorded_seals)
    if not record.seal_check_passed:
        mismatch = f'⚠️ عدم تطابق الأختام | المسجّل: {trip.seal_numbers} | المشاهَد: {observed_input}'
        record.seal_notes = f'{record.seal_notes} | {mismatch}' if record.seal_notes else mismatch

    record.save()

    if record.is_suspect:
        trip.status = 'suspect'
        # سبب الاشتباه: عجز حجمي و/أو عدم تطابق الأختام
        reasons = []
        if not record.seal_check_passed:
            reasons.append('عدم تطابق أرقام الأختام')
        if record.variance_liters and abs(record.variance_liters) > 0:
            reasons.append(f'عجز حجمي {record.variance_liters} لتر ({record.variance_percent}%)')
        reason_text = ' + '.join(reasons) if reasons else 'مخالفة في القراءات'
        Alert.objects.create(
            trip=trip, alert_type='volume_variance', severity='critical',
            description=f'شحنة مشبوهة عند تفريغ الرحلة {trip.trip_code}: {reason_text}.',
        )
        messages.warning(request, f'تم حفظ السجل، لكن رُصدت مخالفة ({reason_text}). صُنّفت الرحلة كمشبوهة وأُخطر المراقبون.')
    else:
        trip.status = 'completed'
        messages.success(request, 'تم تسجيل التفريغ والتحقق من المطابقة الرقمية والجغرافية والأختام بنجاح. أُغلقت الرحلة كمكتملة.')

    trip.save(update_fields=['status', 'updated_at'])
    AuditLog.objects.create(
        user=request.user, action='submit_discharge',
        target_table='shield_trips', target_id=trip.id, ip_address=ip,
        new_value=f'الحالة النهائية: {trip.get_status_display()}',
    )
    return redirect('trips:discharge_log')



@login_required
def report_fault(request, trip_id):
    """معالجة بلاغات فروقات القراءات أو أعطال الأجهزة، مربوطة بالمحطة والرحلة الحالية."""
    if not _require_agent(request):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة.')
        return redirect('dashboard:home')

    trip = get_object_or_404(Trip, pk=trip_id, agent=request.user)
    ip = _get_client_ip(request)

    if not is_trusted_device(request):
        AuditLog.objects.create(
            user=request.user, action='untrusted_device_breach',
            target_table='shield_mechanical_faults', target_id=trip.id, ip_address=ip,
            new_value=f'محاولة رفع بلاغ للرحلة {trip.trip_code} من جهاز غير موثّق.',
        )
        Alert.objects.create(
            trip=trip, alert_type='unauthorized_device', severity='critical',
            description=f'محاولة رفع بلاغ للرحلة {trip.trip_code} من جهاز غير معتمد للمستخدم {request.user.username}. رُفضت العملية.',
        )
        messages.error(request, 'تم رفض العملية: هذا الجهاز غير موثوق أمنياً. تواصل مع مدير النظام لاعتماد جهازك.')
        return redirect('trips:agent_trip_list')

    if request.method == 'POST':
        form = FaultReportForm(request.POST, request.FILES)
        if form.is_valid():
            fault = form.save(commit=False)
            fault.reported_by = request.user
            fault.trip = trip
            fault.station = trip.station
            fault.truck = trip.truck
            fault.save()
            AuditLog.objects.create(
                user=request.user, action='report_fault',
                target_table='shield_mechanical_faults', target_id=fault.id, ip_address=ip,
                new_value=f'{fault.get_fault_type_display()} - الرحلة {trip.trip_code}',
            )
            messages.success(request, 'تم تسجيل بلاغك بنجاح وسيتم مراجعته من الفريق الفني قريباً.')
            return redirect('trips:agent_fault_list')
    else:
        form = FaultReportForm()

    return render(request, 'trips/fault_report.html', {'trip': trip, 'form': form})
@user_passes_test(_is_admin, login_url='users:login')
def driver_list(request):
    qs = Driver.objects.select_related('warehouse').order_by('driver_name')
    search = request.GET.get('q', '').strip()
    if search:
        qs = qs.filter(Q(driver_name__icontains=search) | Q(national_id__icontains=search) | Q(license_number__icontains=search))
    return render(request, 'trips/driver_list.html', {'drivers': qs, 'search': search})


@user_passes_test(_is_admin, login_url='users:login')
def driver_create(request):
    if request.method == 'POST':
        form = DriverForm(request.POST)
        if form.is_valid():
            d = form.save()
            AuditLog.objects.create(user=request.user, action='create_driver', target_table='drivers',
                                    target_id=d.id, new_value=f'{d.driver_name} - {d.license_number}',
                                    ip_address=get_client_ip(request))
            messages.success(request, f'✅ تم إضافة السائق «{d.driver_name}» بنجاح.')
            return redirect('trips:driver_list')
    else:
        form = DriverForm()
    return render(request, 'trips/driver_form.html', {'form': form, 'mode': 'create'})


@user_passes_test(_is_admin, login_url='users:login')
def driver_update(request, pk):
    d = get_object_or_404(Driver, pk=pk)
    if request.method == 'POST':
        form = DriverForm(request.POST, instance=d)
        if form.is_valid():
            d = form.save()
            AuditLog.objects.create(user=request.user, action='update_driver', target_table='drivers',
                                    target_id=d.id, new_value=f'{d.driver_name} - {d.license_number}',
                                    ip_address=get_client_ip(request))
            messages.success(request, f'✅ تم تحديث بيانات السائق «{d.driver_name}».')
            return redirect('trips:driver_list')
    else:
        form = DriverForm(instance=d)
    return render(request, 'trips/driver_form.html', {'form': form, 'mode': 'edit', 'driver': d})


@user_passes_test(_is_admin, login_url='users:login')
@require_http_methods(["POST"])
def driver_archive(request, pk):
    d = get_object_or_404(Driver, pk=pk)
    d.is_active = not d.is_active
    d.save(update_fields=['is_active'])
    AuditLog.objects.create(user=request.user,
                            action='activate_driver' if d.is_active else 'deactivate_driver',
                            target_table='drivers', target_id=d.id,
                            new_value='active' if d.is_active else 'inactive',
                            ip_address=get_client_ip(request))
    messages.success(request, f'✅ تم {"تفعيل" if d.is_active else "إيقاف"} السائق «{d.driver_name}».')
    return redirect('trips:driver_list')
@user_passes_test(_is_admin, login_url='users:login')
def truck_create(request):
    if request.method == 'POST':
        form = TruckForm(request.POST)
        if form.is_valid():
            t = form.save()
            AuditLog.objects.create(user=request.user, action='create_truck', target_table='shield_trucks',
                                    target_id=t.id, new_value=f'{t.truck_id} - {t.driver.driver_name}',
                                    ip_address=get_client_ip(request))
            messages.success(request, f'✅ تم إضافة الشاحنة «{t.truck_id}» بنجاح.')
            return redirect('trips:truck_list')
    else:
        form = TruckForm()
    return render(request, 'trips/truck_form.html', {'form': form, 'mode': 'create'})


@user_passes_test(_is_admin, login_url='users:login')
def truck_update(request, pk):
    t = get_object_or_404(Truck, pk=pk)
    if request.method == 'POST':
        form = TruckForm(request.POST, instance=t)
        if form.is_valid():
            t = form.save()
            AuditLog.objects.create(user=request.user, action='update_truck', target_table='shield_trucks',
                                    target_id=t.id, new_value=f'{t.truck_id} - {t.get_status_display()}',
                                    ip_address=get_client_ip(request))
            messages.success(request, f'✅ تم تحديث بيانات الشاحنة «{t.truck_id}».')
            return redirect('trips:truck_list')
    else:
        form = TruckForm(instance=t)
    return render(request, 'trips/truck_form.html', {'form': form, 'mode': 'edit', 'truck': t})
def truck_archive(request, pk):
    t = get_object_or_404(Truck, pk=pk)
    t.is_active = not t.is_active
    t.save(update_fields=['is_active'])
    AuditLog.objects.create(user=request.user,
                            action='activate_truck' if t.is_active else 'deactivate_truck',
                            target_table='shield_trucks', target_id=t.id,
                            new_value='active' if t.is_active else 'inactive',
                            ip_address=get_client_ip(request))
    messages.success(request, f'✅ تم {"تفعيل" if t.is_active else "إيقاف"} الشاحنة «{t.truck_id}».')
    return redirect('trips:truck_list')
#طباعة QR
def trip_qr(request, trip_id):
    """توليد وعرض رمز QR للرحلة للطباعة (موظف المستودع)."""
    if request.user.role not in ('dispatcher', 'manager', 'admin'):
        messages.error(request, 'ليس لديك صلاحية.')
        return redirect('dashboard:home')

    trip = get_object_or_404(
        Trip.objects.select_related('station', 'truck', 'truck__driver'), pk=trip_id
    )
    if request.user.role == 'dispatcher' and trip.dispatcher != request.user:
        messages.error(request, 'هذه الرحلة ليست من تجهيزك.')
        return redirect('trips:trip_list')

    # ===== التحقق من اكتمال البيانات المرجعية (المجرى البديل) =====
    missing = []
    if not trip.station_id:  missing.append('المحطة')
    if not trip.truck_id:    missing.append('الشاحنة')
    if not trip.seal_numbers: missing.append('أرقام الأختام')
    if not trip.qr_token:    missing.append('توكن QR')
    if missing:
        messages.error(request, 'تعذّر توليد الرمز — بيانات ناقصة: ' + '، '.join(missing) + '.')
        return redirect('trips:trip_detail', trip_id=trip.id)

    qr_url = request.build_absolute_uri(reverse('trips:qr_redirect', args=[trip.qr_token]))
    return render(request, 'trips/trip_qr.html', {'trip': trip, 'qr_url': qr_url})

