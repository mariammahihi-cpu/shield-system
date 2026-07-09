from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q
from users.decorators import role_required
from .models import AuditLog, AdminAction
from audit.utils import record_audit, record_admin_action
from django.http import HttpResponse
from django.utils import timezone
from django.shortcuts import redirect
from django.contrib import messages



# إجراءات تُعتبر حسّاسة/مشبوهة (تُميَّز بصرياً)
SUSPICIOUS_ACTIONS = {
    'qr_mismatch_breach', 'auto_lock', 'geofence_violation',
    'deactivate_driver', 'deactivate_truck', 'untrust_device', 'seal_tampering',
}



@role_required('admin')
def audit_log_list(request):
    if request.user.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة.')
        return redirect('dashboard:home')
    request.user.audit_seen_at = timezone.now()
    request.user.save(update_fields=['audit_seen_at'])
    """سجل التدقيق العام مع بحث وفلترة وتصدير."""
    qs = AuditLog.objects.select_related('user').order_by('-timestamp')

    search    = request.GET.get('q', '').strip()
    table     = request.GET.get('table', '').strip()
    date_from = request.GET.get('from', '').strip()
    date_to   = request.GET.get('to', '').strip()

    if search:
        qs = qs.filter(Q(action__icontains=search) | Q(user__username__icontains=search))
    if table:
        qs = qs.filter(target_table=table)
    if date_from:
        qs = qs.filter(timestamp__date__gte=date_from)
    if date_to:
        qs = qs.filter(timestamp__date__lte=date_to)

    if request.GET.get('export') == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'سجل التدقيق'
        ws.sheet_view.rightToLeft = True

        headers = ['الإجراء', 'الجدول', 'المعرّف', 'المستخدم', 'IP', 'القيمة القديمة', 'القيمة الجديدة', 'الوقت']
        ws.append(headers)
        fill = PatternFill('solid', fgColor='1E3A8A')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for log in qs:
            ws.append([
                log.action, log.target_table, log.target_id or '',
                (log.user.full_name or log.user.username) if log.user else 'نظام',
                log.ip_address or '', log.old_value or '', log.new_value or '',
                log.timestamp.strftime('%Y-%m-%d %H:%M'),
            ])

        for i, w in enumerate([22, 20, 10, 20, 14, 32, 32, 18], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="audit_log_{timezone.now():%Y%m%d_%H%M}.xlsx"'
        wb.save(response)
        return response

    page = Paginator(qs, 25).get_page(request.GET.get('page'))
    tables = AuditLog.objects.values_list('target_table', flat=True).distinct()

    return render(request, 'audit/audit_log_list.html', {
        'page_obj': page, 'search': search, 'table_filter': table,
        'date_from': date_from, 'date_to': date_to, 'tables': tables,
        'suspicious_actions': SUSPICIOUS_ACTIONS,
    })


@role_required('admin')
def admin_action_list(request):
    if request.user.role != 'admin':
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة.')
        return redirect('dashboard:home')
    """سجل الإجراءات الإدارية الحسّاسة مع بحث وفلترة وتصدير."""
    qs = AdminAction.objects.select_related('target_user', 'admin_user').order_by('-created_at')

    action    = request.GET.get('action', '').strip()
    search    = request.GET.get('q', '').strip()
    date_from = request.GET.get('from', '').strip()
    date_to   = request.GET.get('to', '').strip()

    if action:
        qs = qs.filter(action_type=action)
    if search:
        qs = qs.filter(
            Q(target_user__username__icontains=search) |
            Q(target_user__full_name__icontains=search) |
            Q(admin_user__username__icontains=search) |
            Q(reason__icontains=search)
        )
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    # ===== تصدير Excel احترافي =====
    if request.GET.get('export') == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook(); ws = wb.active
        ws.title = 'الإجراءات الإدارية'
        ws.sheet_view.rightToLeft = True

        headers = ['الإجراء', 'المستخدم المستهدف', 'المنفِّذ', 'المبرّر', 'الحالة', 'الوقت']
        ws.append(headers)
        fill = PatternFill('solid', fgColor='1E3A8A')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for a in qs:
            ws.append([
                a.get_action_type_display(),
                a.target_user.full_name or a.target_user.username,
                (a.admin_user.full_name or a.admin_user.username) if a.admin_user else 'النظام (تلقائي)',
                a.reason,
                'سارٍ' if a.is_active else 'منتهٍ',
                a.created_at.strftime('%Y-%m-%d %H:%M'),
            ])

        for i, w in enumerate([24, 22, 20, 40, 12, 18], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="admin_actions_{timezone.now():%Y%m%d_%H%M}.xlsx"'
        wb.save(response)
        return response

    page = Paginator(qs, 25).get_page(request.GET.get('page'))
    return render(request, 'audit/admin_action_list.html', {
        'page_obj': page, 'action_filter': action,
        'search': search, 'date_from': date_from, 'date_to': date_to,
        'action_choices': AdminAction.ACTION_CHOICES,
    })